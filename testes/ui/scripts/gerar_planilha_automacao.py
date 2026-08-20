"""
Gera a planilha de automacao (Excel) do projeto SIGNA a partir do estado real do
repositorio: varre os arquivos .feature de cypress/e2e/api e cypress/e2e/ui,
conta cenarios, detecta fluxos de excecao, verbos HTTP / URLs e datas reais de
git, estima horas e classifica cada funcionalidade nas sprints 04-13.

Uso:
    python scripts/gerar_planilha_automacao.py

Saida:
    cypress/e2e/api/planilha_automacao.xlsx
"""

import re
import subprocess
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
API_DIR = BASE_DIR / "cypress" / "e2e" / "api"
UI_DIR = BASE_DIR / "cypress" / "e2e" / "ui"
OUTPUT_PATH = API_DIR / "planilha_automacao.xlsx"

QA_NOME = "Marcus Vinicius Silva da Rocha"

# ---------------------------------------------------------------------------
# Parametros de estimativa e capacidade (ajustar aqui conforme o time real)
# ---------------------------------------------------------------------------
SETUP_HORAS_API = 3
HORAS_POR_CENARIO_API = 1
SETUP_HORAS_UI = 5
HORAS_POR_CENARIO_UI = 2
CAPACIDADE_HORAS_POR_SPRINT = 60  # 1 QA x 6h/dia util x ~10 dias uteis

SPRINT_INICIAL = 4
SPRINT_FINAL = 13
JANELA_INICIO = date(2026, 1, 20)
JANELA_FIM = date(2026, 6, 12)

# Multiplicador de complexidade para telas UI com fluxos maiores que um
# cenario simples costuma cobrir (wizards, varias abas, varios sub-fluxos).
COMPLEXIDADE_UI = {
    "designacao": 1.15,
    "cessacao": 1.6,
    "apostilar": 1.73,
    "insubsistente": 2.08,
    # consulta_atos_adminstra (ex-editar_designa_ao) absorveu o antigo
    # visualiza_designação.feature: os dois fluxos convergem para a mesma
    # tela (visualizar-designacao/{id}) desde a migração do menu lateral,
    # então viraram 2 cenários num só arquivo em vez de dois quase idênticos.
    "consulta_atos_adminstra": 1.57,
    "altera_DO": 1.21,
}

# Quando a feature ainda e um stub (0 cenarios escritos) usamos uma
# quantidade planejada de cenarios para nao subestimar o esforco restante.
CENARIOS_PLANEJADOS_STUB = {}

CRUD_UI = {
    "login": "R",
    "esqueci_senha": "U",
    "alterar_senha": "U",
    "alteracao_email": "U",
    "designacao": "C",
    "consulta_atos_adminstra": "U/R",
    "cessacao": "U",
    "apostilar": "U",
    "insubsistente": "U",
    "altera_DO": "U",
}

# Funcionalidades de API cujo POST e usado como filtro/consulta (nao "Create")
FORCA_CRUD_R = {"api_cenarios_negativos", "api_diretoria_regional"}

# feature (sem extensao) -> step_definitions correspondente, para extrair a URL real
FEATURE_TO_STEPFILE = {
    "login": "login_steps.js",
    "esqueci_senha": "esqueci_senha_steps.js",
    "alterar_senha": "alterar_senha_steps.js",
    "alteracao_email": "alteracao_email_steps.js",
    "designacao": "designacao_steps.js",
    "consulta_atos_adminstra": "editar_designacao_steps.js",
    "cessacao": "cessacao_steps.js",
    "apostilar": "apostilar_steps.js",
    "insubsistente": "insubsistente_steps.js",
    "altera_DO": "altera_DO_steps.js",
}

# Fallback manual quando nao ha cy.visit/should(include) explicito no codigo
# (fluxos que navegam via clique a partir da listagem, sem URL fixa propria).
URL_FALLBACK_UI = {
    "login": "/login",
    "insubsistente": "/pages/listagem-designacoes (acao Insubsistente)",
    "apostilar": "/pages/listagem-designacoes (acao Apostilar)",
    "consulta_atos_adminstra": "/pages/listagem-designacoes/visualizar-designacao/{id} (acoes Editar e Detalhar)",
    "alterar_senha": "/meus-dados (modal Alterar Senha)",
}

# Fallback para features de API cujo texto de comentarios nao segue o padrao
# "# VERBO /caminho" (ex.: suite cruzada de cenarios negativos).
URL_FALLBACK_API = {
    "api_cenarios_negativos": "Multiplos endpoints (ver /api/DREs, /api/abrangencia/*, /api/acessos/*)",
}

# Nome legivel para features que ainda sao apenas um stub (sem linha
# "Funcionalidade:" escrita).
FUNCIONALIDADE_FALLBACK = {}

PALAVRAS_EXCECAO = re.compile(
    r"negativ|inv[aá]lid|inexistente|sem (login|token|auth)|erro|obrigat[oó]rio|"
    r"vazio|sql|injection|traversal|n[aã]o deve|excede",
    re.IGNORECASE,
)

VERBO_CRUD = {"GET": "R", "POST": "C", "PUT": "U", "PATCH": "U", "DELETE": "D"}


@dataclass
class Linha:
    sistema: str
    funcionalidade: str
    crud: str
    url: str
    qtd_cenarios: int
    qtd_excecao: int
    horas_api: int
    horas_ui: int
    finalizado: str
    data_inicio: date | None
    data_fim: date | None
    sprint: str = field(default="")

    @property
    def horas_total(self) -> int:
        return self.horas_api + self.horas_ui


def git(*args: str) -> str:
    result = subprocess.run(
        ["git", *args], cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8"
    )
    return result.stdout.strip()


def is_tracked(path: Path) -> bool:
    rel = path.relative_to(BASE_DIR).as_posix()
    result = subprocess.run(
        ["git", "ls-files", "--error-unmatch", rel],
        cwd=BASE_DIR,
        capture_output=True,
        text=True,
    )
    return result.returncode == 0


def git_dates(path: Path) -> tuple[date | None, date | None]:
    """Data de inicio = primeiro commit que criou o arquivo.
    Data de fim = commit mais recente que alterou conteudo de cenario
    (evita datar como "finalizacao" commits cosmeticos, ex.: ajuste de tag
    ou .gitignore, que tambem tocam o arquivo).
    """
    rel = path.relative_to(BASE_DIR).as_posix()
    log = git("log", "--follow", "--format=%H %ad", "--date=short", "--", rel)
    linhas = log.splitlines()
    if not linhas:
        return None, None
    commits = [linha.split(" ", 1) for linha in linhas]  # mais novo -> mais velho
    inicio = date.fromisoformat(commits[-1][1])
    fim = date.fromisoformat(commits[0][1])
    for sha, quando in commits:
        diff = git("show", sha, "--", rel)
        if re.search(r"^[+-]\s*Cen[aá]rio", diff, re.MULTILINE):
            fim = date.fromisoformat(quando)
            break
    return inicio, fim


def sprint_windows() -> list[tuple[int, date, date]]:
    total_dias = (JANELA_FIM - JANELA_INICIO).days + 1
    n = SPRINT_FINAL - SPRINT_INICIAL + 1
    janelas = []
    cursor = JANELA_INICIO
    restante_dias, restante_sprints = total_dias, n
    for i in range(n):
        tamanho = -(-restante_dias // restante_sprints)  # ceil
        fim = cursor + timedelta(days=tamanho - 1)
        janelas.append((SPRINT_INICIAL + i, cursor, fim))
        cursor = fim + timedelta(days=1)
        restante_dias -= tamanho
        restante_sprints -= 1
    return janelas


SPRINTS = sprint_windows()


def rotulo_sprint(d: date | None) -> str:
    if d is None:
        return "Em andamento / sem data"
    for numero, ini, fim in SPRINTS:
        if ini <= d <= fim:
            return f"Sprint {numero:02d}"
    if d < JANELA_INICIO:
        return "Anterior a Sprint 04"
    return "Apos Sprint 13 (fora da janela)"


def contar_cenarios(texto: str) -> tuple[int, int]:
    titulos = re.findall(r"^\s*Cen[aá]rio(?:\s+de\s+Fundo)?:\s*(.+)$", texto, re.MULTILINE)
    total = len(titulos)
    excecao = sum(1 for t in titulos if PALAVRAS_EXCECAO.search(t))
    return total, excecao


def nome_funcionalidade(stem: str, texto: str) -> str:
    m = re.search(r"^Funcionalidade:\s*(.+)$", texto, re.MULTILINE)
    if m:
        return re.sub(r"^API EOL - ", "", m.group(1).strip())
    return FUNCIONALIDADE_FALLBACK.get(stem, stem)


def extrair_urls_api(stem: str, texto: str) -> tuple[str, set[str]]:
    pares = re.findall(r"#\s+(GET|POST|PUT|PATCH|DELETE)\s+(/\S+)", texto)
    if not pares:
        pares = re.findall(
            r"Quero validar o endpoint (GET|POST|PUT|PATCH|DELETE)\s+(/\S+)", texto
        )
    verbos_reais = set(re.findall(r"requisi[cç][aã]o (GET|POST|PUT|PATCH|DELETE)", texto))
    urls = " | ".join(dict.fromkeys(f"{verbo} {caminho}" for verbo, caminho in pares))
    if not urls:
        urls = URL_FALLBACK_API.get(stem, "A definir")
    return urls, (verbos_reais or {v for v, _ in pares})


def _urls_em(path: Path) -> list[str]:
    if not path.exists():
        return []
    texto = path.read_text(encoding="utf-8")
    achados = re.findall(r"cy\.visit\('([^']+)'\)", texto)
    achados += re.findall(r"should\('include',\s*'([^']+)'\)", texto)
    return [u for u in achados if u not in ("/", "")]


def extrair_url_ui(stem: str) -> str:
    # Nao usamos os commands globais (ex.: commands_login.js) como fonte:
    # eles sao compartilhados por praticamente todo fluxo (login no Contexto)
    # e poluiriam qualquer funcionalidade sem cy.visit proprio com "/login".
    step_file = FEATURE_TO_STEPFILE.get(stem)
    achados = _urls_em(BASE_DIR / "cypress" / "support" / "step_definitions" / "ui" / step_file) if step_file else []

    unicos = list(dict.fromkeys(achados))
    if unicos:
        return " | ".join(unicos)
    return URL_FALLBACK_UI.get(stem, "A definir")


def processar_api(path: Path) -> Linha:
    texto = path.read_text(encoding="utf-8")
    stem = path.stem
    funcionalidade = nome_funcionalidade(stem, texto)
    qtd_cenarios, qtd_excecao = contar_cenarios(texto)
    url, verbos = extrair_urls_api(stem, texto)

    if stem in FORCA_CRUD_R:
        crud = "R"
    else:
        letras = sorted({VERBO_CRUD[v] for v in verbos if v in VERBO_CRUD})
        crud = "".join(letras) or "R"

    horas_api = SETUP_HORAS_API + qtd_cenarios * HORAS_POR_CENARIO_API

    tracked = is_tracked(path)
    inicio, fim = git_dates(path) if tracked else (None, None)
    finalizado = "Sim" if tracked else ("Em andamento" if qtd_cenarios else "Nao iniciado")

    return Linha(
        sistema="EOL (Integracao SGP)",
        funcionalidade=funcionalidade,
        crud=crud,
        url=url,
        qtd_cenarios=qtd_cenarios,
        qtd_excecao=qtd_excecao,
        horas_api=horas_api,
        horas_ui=0,
        finalizado=finalizado,
        data_inicio=inicio,
        data_fim=fim,
    )


def processar_ui(path: Path) -> Linha:
    texto = path.read_text(encoding="utf-8")
    stem = path.stem
    funcionalidade = nome_funcionalidade(stem, texto)
    qtd_cenarios, qtd_excecao = contar_cenarios(texto)

    efetivos = qtd_cenarios or CENARIOS_PLANEJADOS_STUB.get(stem, 0)
    complexidade = COMPLEXIDADE_UI.get(stem, 1.0)
    horas_ui = round((SETUP_HORAS_UI + efetivos * HORAS_POR_CENARIO_UI) * complexidade)

    crud = CRUD_UI.get(stem, "A definir")
    url = extrair_url_ui(stem)

    tracked = is_tracked(path)
    inicio, fim = git_dates(path) if tracked else (None, None)
    if tracked:
        finalizado = "Sim"
    elif qtd_cenarios:
        finalizado = "Em andamento"
    else:
        finalizado = "Nao iniciado"

    return Linha(
        sistema="SIGNA",
        funcionalidade=funcionalidade,
        crud=crud,
        url=url,
        qtd_cenarios=qtd_cenarios,
        qtd_excecao=qtd_excecao,
        horas_api=0,
        horas_ui=horas_ui,
        finalizado=finalizado,
        data_inicio=inicio,
        data_fim=fim,
    )


def coletar_linhas() -> list[Linha]:
    linhas = [processar_ui(p) for p in sorted(UI_DIR.glob("*.feature"))]
    linhas += [processar_api(p) for p in sorted(API_DIR.glob("*.feature"))]
    for linha in linhas:
        linha.sprint = rotulo_sprint(linha.data_fim)
    return linhas


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="C6E0B4")
HEADER_FONT = Font(bold=True)
STATUS_FILL = {
    "Sim": PatternFill("solid", fgColor="C6E0B4"),
    "Em andamento": PatternFill("solid", fgColor="FFE699"),
    "Nao iniciado": PatternFill("solid", fgColor="F8CBAD"),
}

COLUNAS = [
    ("Sistema", "sistema", 16),
    ("QA que Iniciou a Automacao", None, 24),
    ("Funcionalidade", "funcionalidade", 42),
    ("CRUD", "crud", 10),
    ("URL", "url", 46),
    ("Quantidade de Cenarios", "qtd_cenarios", 12),
    ("Fluxo de Excecao", "qtd_excecao", 10),
    ("Horas API (Endpoint)", "horas_api", 12),
    ("Horas UI (Telas)", "horas_ui", 12),
    ("Horas total", "horas_total", 10),
    ("Finalizado", "finalizado", 14),
    ("Data Inicio da Automacao", "data_inicio", 16),
    ("Data Finalizacao da Automacao", "data_fim", 18),
    ("QA que Finalizou a Automacao", None, 24),
]


def montar_planilha_principal(wb: Workbook, linhas: list[Linha]) -> None:
    ws = wb.active
    ws.title = "Planilha Automacao"

    for col, (titulo, _, largura) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = largura

    for row, linha in enumerate(linhas, start=2):
        valores = [
            linha.sistema,
            QA_NOME,
            linha.funcionalidade,
            linha.crud,
            linha.url,
            linha.qtd_cenarios,
            linha.qtd_excecao,
            linha.horas_api,
            linha.horas_ui,
            linha.horas_total,
            linha.finalizado,
            linha.data_inicio.strftime("%d/%m/%Y") if linha.data_inicio else ("Em andamento" if linha.finalizado != "Nao iniciado" else "A iniciar"),
            linha.data_fim.strftime("%d/%m/%Y") if linha.data_fim else "Em andamento",
            QA_NOME if linha.finalizado == "Sim" else "",
        ]
        for col, valor in enumerate(valores, start=1):
            ws.cell(row=row, column=col, value=valor)
        ws.cell(row=row, column=11).fill = STATUS_FILL.get(linha.finalizado, PatternFill())

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{len(linhas) + 1}"


def montar_resumo_sprint(wb: Workbook, linhas: list[Linha]) -> None:
    ws = wb.create_sheet("Resumo por Sprint")
    cabecalho = ["Sprint", "Periodo", "Capacidade (h)", "Horas Entregues", "Entregas"]
    for col, titulo in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    por_sprint: dict[str, list[Linha]] = {}
    for linha in linhas:
        por_sprint.setdefault(linha.sprint, []).append(linha)

    row = 2
    total_estimado = 0
    for numero, ini, fim in SPRINTS:
        label = f"Sprint {numero:02d}"
        entregas = por_sprint.get(label, [])
        horas = sum(l.horas_total for l in entregas)
        total_estimado += horas
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=f"{ini.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}")
        ws.cell(row=row, column=3, value=CAPACIDADE_HORAS_POR_SPRINT)
        ws.cell(row=row, column=4, value=horas)
        ws.cell(row=row, column=5, value=", ".join(l.funcionalidade for l in entregas) or "-")
        row += 1

    outros_labels = set(por_sprint) - {f"Sprint {n:02d}" for n, _, _ in SPRINTS}
    for label in sorted(outros_labels):
        entregas = por_sprint[label]
        horas = sum(l.horas_total for l in entregas)
        total_estimado += horas
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value="-")
        ws.cell(row=row, column=3, value="-")
        ws.cell(row=row, column=4, value=horas)
        ws.cell(row=row, column=5, value=", ".join(l.funcionalidade for l in entregas) or "-")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=4, value=total_estimado)
    ws.cell(row=row, column=5, value=f"Capacidade total do periodo: {CAPACIDADE_HORAS_POR_SPRINT * (SPRINT_FINAL - SPRINT_INICIAL + 1)}h")

    for col, largura in enumerate([12, 24, 14, 16, 90], start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura


def main() -> None:
    linhas = coletar_linhas()
    wb = Workbook()
    montar_planilha_principal(wb, linhas)
    montar_resumo_sprint(wb, linhas)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Planilha gerada em: {OUTPUT_PATH}")
    print(f"{len(linhas)} funcionalidades processadas.")


if __name__ == "__main__":
    main()
